import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar

class ImageEnhancementReportGenerator:
    def __init__(self, csv_file_path):
        """Initialize the report generator with the CSV file path."""
        self.csv_file_path = csv_file_path
        self.df = None
        self.load_and_clean_data()
    
    def load_and_clean_data(self):
        """Load and clean the CSV data."""
        try:
            # Read CSV with robust parsing
            self.df = pd.read_csv(self.csv_file_path, encoding='utf-8-sig')
            
            # Clean column names by stripping whitespace
            self.df.columns = self.df.columns.str.strip()
            
            # Remove unnamed columns
            self.df = self.df.loc[:, ~self.df.columns.str.contains('^Unnamed')]
            
            # Convert workdate to datetime
            self.df['workdate'] = pd.to_datetime(self.df['workdate'], format='%d/%m/%Y')
            
            # Fill NaN values with 0 for numeric columns
            numeric_columns = ['TotalDone', 'Good', 'GoodOriginal', 'GoodEnhanced', 
                             'ForDownload', 'Bad', 'Rejected', 'Downloaded', 'Uploaded']
            
            for col in numeric_columns:
                if col in self.df.columns:
                    self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0)
            
            # Add calculated columns
            self.df['ForEditing'] = self.df['GoodEnhanced'] + self.df['ForDownload']
            self.df['TotalReviewed'] = self.df['Good'] + self.df['Bad'] + self.df['Rejected']
            self.df['TotalEdited'] = self.df['Downloaded'] + self.df['Uploaded']
            
            print(f"Data loaded successfully. Shape: {self.df.shape}")
            
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            raise
    
    def get_months_in_data(self):
        """Get all unique months present in the data."""
        if self.df is None or self.df.empty:
            return []
        
        months = self.df['workdate'].dt.to_period('M').unique()
        return sorted(months)
    
    def get_weeks_in_month(self, year_month):
        """Get all weeks that belong to a specific month (Monday to Friday weeks)."""
        # Get the first and last day of the month
        month_start = year_month.start_time
        month_end = year_month.end_time
        
        # Find all weeks that have any day in this month
        all_data = self.df.copy()
        all_data['week_period'] = all_data['workdate'].dt.to_period('W')
        
        # Get all weeks where the week overlaps with the month
        weeks_in_month = []
        for week_period in all_data['week_period'].unique():
            week_start = week_period.start_time  # Monday
            week_end = week_period.end_time      # Sunday
            
            # Check if this week overlaps with the month
            # A week belongs to a month if any of its weekdays (Mon-Fri) fall in that month
            week_weekdays_start = week_start
            week_weekdays_end = week_start + timedelta(days=4)  # Friday
            
            if (week_weekdays_start <= month_end and week_weekdays_end >= month_start):
                weeks_in_month.append(week_period)
        
        return sorted(weeks_in_month)
    
    def filter_data_by_month(self, year_month):
        """Filter data for a specific month."""
        month_data = self.df[self.df['workdate'].dt.to_period('M') == year_month].copy()
        return month_data.sort_values(['workdate', 'useremail'])
    
    def filter_data_by_week(self, week_period):
        """Filter data for a specific week (Monday to Friday only)."""
        week_start = week_period.start_time  # Monday
        week_end = week_start + timedelta(days=4)  # Friday
        
        week_data = self.df[
            (self.df['workdate'] >= week_start) & 
            (self.df['workdate'] <= week_end) &
            (self.df['workdate'].dt.dayofweek < 5)  # Monday=0 to Friday=4
        ].copy()
        return week_data.sort_values(['workdate', 'useremail'])
    
    def get_weekday_name(self, date):
        """Get the weekday name for a given date."""
        return date.strftime('%A')
    
    def generate_daily_summary(self, data, is_week_sheet=False):
        """Generate daily summary with totals."""
        daily_summary = []
        
        # Group by date
        for date, day_data in data.groupby('workdate'):
            weekday = self.get_weekday_name(date)
            date_str = date.strftime('%d/%m/%Y')
            
            # Add individual user records for this day
            for _, row in day_data.iterrows():
                daily_summary.append({
                    'WorkDate': date_str,
                    'Weekday': weekday,
                    'User': row['useremail'],
                    'TotalDone': row['TotalDone'],
                    'Good': row['Good'],
                    'GoodOriginal': row['GoodOriginal'],
                    'GoodEnhanced': row['GoodEnhanced'],
                    'ForEditing': row['ForEditing'],
                    'Bad': row['Bad'],
                    'Rejected': row['Rejected'],
                    'TotalReviewed': row['TotalReviewed'],
                    'Downloaded': row['Downloaded'],
                    'Uploaded': row['Uploaded'],
                    'TotalEdited': row['TotalEdited']
                })
            
            # Add daily total for both week and month sheets
            daily_total = {
                'WorkDate': f'TOTAL {weekday.upper()}',
                'Weekday': '',
                'User': '',
                'TotalDone': day_data['TotalDone'].sum(),
                'Good': day_data['Good'].sum(),
                'GoodOriginal': day_data['GoodOriginal'].sum(),
                'GoodEnhanced': day_data['GoodEnhanced'].sum(),
                'ForEditing': day_data['ForEditing'].sum(),
                'Bad': day_data['Bad'].sum(),
                'Rejected': day_data['Rejected'].sum(),
                'TotalReviewed': day_data['TotalReviewed'].sum(),
                'Downloaded': day_data['Downloaded'].sum(),
                'Uploaded': day_data['Uploaded'].sum(),
                'TotalEdited': day_data['TotalEdited'].sum()
            }
            daily_summary.append(daily_total)
        
        # Add period total (weekly total for week sheets, monthly total for month sheets)
        if not data.empty:
            if is_week_sheet:
                period_total_name = 'WEEKLY TOTAL'
            else:
                period_total_name = 'MONTHLY TOTAL'
                
            period_total = {
                'WorkDate': period_total_name,
                'Weekday': '',
                'User': '',
                'TotalDone': data['TotalDone'].sum(),
                'Good': data['Good'].sum(),
                'GoodOriginal': data['GoodOriginal'].sum(),
                'GoodEnhanced': data['GoodEnhanced'].sum(),
                'ForEditing': data['ForEditing'].sum(),
                'Bad': data['Bad'].sum(),
                'Rejected': data['Rejected'].sum(),
                'TotalReviewed': data['TotalReviewed'].sum(),
                'Downloaded': data['Downloaded'].sum(),
                'Uploaded': data['Uploaded'].sum(),
                'TotalEdited': data['TotalEdited'].sum()
            }
            daily_summary.append(period_total)
        
        return daily_summary
    
    def generate_user_summary(self, data, period_name):
        """Generate summary by user."""
        user_summary = []
        
        for user, user_data in data.groupby('useremail'):
            summary = {
                'User': user,
                'TotalDone': user_data['TotalDone'].sum(),
                'Good': user_data['Good'].sum(),
                'GoodOriginal': user_data['GoodOriginal'].sum(),
                'GoodEnhanced': user_data['GoodEnhanced'].sum(),
                'ForEditing': user_data['ForEditing'].sum(),
                'Bad': user_data['Bad'].sum(),
                'Rejected': user_data['Rejected'].sum(),
                'TotalReviewed': user_data['TotalReviewed'].sum(),
                'Downloaded': user_data['Downloaded'].sum(),
                'Uploaded': user_data['Uploaded'].sum(),
                'TotalEdited': user_data['TotalEdited'].sum()
            }
            user_summary.append(summary)
        
        # Sort by user email
        user_summary.sort(key=lambda x: x['User'])
        
        # Add period total
        period_total = {
            'User': f'{period_name.upper()} TOTAL',
            'TotalDone': data['TotalDone'].sum(),
            'Good': data['Good'].sum(),
            'GoodOriginal': data['GoodOriginal'].sum(),
            'GoodEnhanced': data['GoodEnhanced'].sum(),
            'ForEditing': data['ForEditing'].sum(),
            'Bad': data['Bad'].sum(),
            'Rejected': data['Rejected'].sum(),
            'TotalReviewed': data['TotalReviewed'].sum(),
            'Downloaded': data['Downloaded'].sum(),
            'Uploaded': data['Uploaded'].sum(),
            'TotalEdited': data['TotalEdited'].sum()
        }
        user_summary.append(period_total)
        
        return user_summary
    
    def create_worksheet(self, wb, sheet_name, daily_summary, user_summary, title, period_dates, is_week_sheet=False):
        """Create a formatted worksheet."""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Clear the worksheet
        ws.delete_rows(1, ws.max_row)
        
        # Set title
        ws['A1'] = title
        ws.merge_cells('A1:O1')
        
        # Style the title
        title_font = Font(bold=True, size=14)
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add section headers for REVIEWER and EDITOR
        ws['D2'] = 'REVIEWER'
        ws.merge_cells('D2:J2')
        ws['D2'].font = Font(bold=True)
        ws['D2'].alignment = Alignment(horizontal='center')
        
        ws['K2'] = 'EDITOR'
        ws.merge_cells('K2:M2')
        ws['K2'].font = Font(bold=True)
        ws['K2'].alignment = Alignment(horizontal='center')
        
        # Add main headers
        headers = ['Work Date', 'User', 'Total Done', 'Good', 'Good Original', 'Good Enhanced', 
                  'For Editing', 'Bad', 'Rejected', 'Total Reviewed', 'Downloaded', 'Uploaded', 'Total Edited']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add daily summary data with merged dates
        row = 4
        current_date = None
        date_start_row = None
        date_rows = {}  # Track which rows belong to each date
        
        for item in daily_summary:
            # Check if this is a new date (exclude TOTAL, WEEKLY TOTAL, and MONTHLY TOTAL rows)
            if (current_date != item['WorkDate'] and 
                not item['WorkDate'].startswith('TOTAL') and 
                not item['WorkDate'].startswith('WEEKLY') and
                not item['WorkDate'].startswith('MONTHLY')):
                
                if current_date is not None and date_start_row is not None:
                    # Store the range for the previous date
                    date_rows[current_date] = (date_start_row, row - 1)
                
                current_date = item['WorkDate']
                date_start_row = row
            
            ws.cell(row=row, column=1, value=item['WorkDate'])
            ws.cell(row=row, column=2, value=item['User'])
            ws.cell(row=row, column=3, value=item['TotalDone'])
            ws.cell(row=row, column=4, value=item['Good'])
            ws.cell(row=row, column=5, value=item['GoodOriginal'])
            ws.cell(row=row, column=6, value=item['GoodEnhanced'])
            ws.cell(row=row, column=7, value=item['ForEditing'])
            ws.cell(row=row, column=8, value=item['Bad'])
            ws.cell(row=row, column=9, value=item['Rejected'])
            ws.cell(row=row, column=10, value=item['TotalReviewed'])
            ws.cell(row=row, column=11, value=item['Downloaded'])
            ws.cell(row=row, column=12, value=item['Uploaded'])
            ws.cell(row=row, column=13, value=item['TotalEdited'])
            
            # Bold the total rows (including TOTAL MONDAY, WEEKLY TOTAL, MONTHLY TOTAL)
            if ('TOTAL' in str(item['WorkDate']) or 'WEEKLY' in str(item['WorkDate']) or 'MONTHLY' in str(item['WorkDate'])):
                for col in range(1, 14):
                    ws.cell(row=row, column=col).font = Font(bold=True)
            
            row += 1
        
        # Store the last date range
        if current_date is not None and date_start_row is not None and row - 1 > date_start_row:
            date_rows[current_date] = (date_start_row, row - 1)
        
        # Now merge and center all date cells (only for actual dates, not TOTAL rows)
        for date_val, (start_row, end_row) in date_rows.items():
            if end_row > start_row and not date_val.startswith('TOTAL'):
                # Merge the cells
                ws.merge_cells(f'A{start_row}:A{end_row}')
                # Center the merged cell both horizontally and vertically
                merged_cell = ws.cell(row=start_row, column=1)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add spacing before summary table - ensure proper alignment
        row += 2
        summary_title = f"Summary Of All Users ({period_dates})"
        
        # Start summary table at column B to align with User column above
        ws.cell(row=row, column=2, value=summary_title)
        ws.merge_cells(f'B{row}:M{row}')  # Span across from B to M
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 2
        
        # Add user summary section headers - properly aligned with data columns
        ws.cell(row=row, column=4, value='REVIEWER')
        ws.merge_cells(f'D{row}:J{row}')
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=11, value='EDITOR')
        ws.merge_cells(f'K{row}:M{row}')
        ws.cell(row=row, column=11).font = Font(bold=True)
        ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
        row += 1
        
        # Add user summary headers - properly aligned starting from column B
        user_headers = ['User', 'Total Done', 'Good', 'Good Original', 'Good Enhanced', 
                       'For Editing', 'Bad', 'Rejected', 'Total Reviewed', 'Downloaded', 'Uploaded', 'Total Edited']
        
        for col, header in enumerate(user_headers, 2):  # Start from column B (2)
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Add user summary data - starting from column B
        for user_data in user_summary:
            ws.cell(row=row, column=2, value=user_data['User'])
            ws.cell(row=row, column=3, value=user_data['TotalDone'])
            ws.cell(row=row, column=4, value=user_data['Good'])
            ws.cell(row=row, column=5, value=user_data['GoodOriginal'])
            ws.cell(row=row, column=6, value=user_data['GoodEnhanced'])
            ws.cell(row=row, column=7, value=user_data['ForEditing'])
            ws.cell(row=row, column=8, value=user_data['Bad'])
            ws.cell(row=row, column=9, value=user_data['Rejected'])
            ws.cell(row=row, column=10, value=user_data['TotalReviewed'])
            ws.cell(row=row, column=11, value=user_data['Downloaded'])
            ws.cell(row=row, column=12, value=user_data['Uploaded'])
            ws.cell(row=row, column=13, value=user_data['TotalEdited'])
            
            # Bold the total row
            if 'TOTAL' in user_data['User']:
                for col in range(2, 14):
                    ws.cell(row=row, column=col).font = Font(bold=True)
            
            row += 1
        
        # Apply borders to all data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = thin_border
        
        # Set appropriate column widths
        from openpyxl.utils import get_column_letter
        
        # Set specific widths for key columns
        column_widths = {
            'A': 12,  # Work Date
            'B': 35,  # User (emails)
            'C': 10,  # Total Done
            'D': 8,   # Good
            'E': 12,  # Good Original
            'F': 12,  # Good Enhanced
            'G': 10,  # For Editing
            'H': 8,   # Bad
            'I': 10,  # Rejected
            'J': 12,  # Total Reviewed
            'K': 12,  # Downloaded
            'L': 10,  # Uploaded
            'M': 12,  # Total Edited
            'N': 12   # Extra column for summary alignment
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    def generate_reports(self, output_directory="reports"):
        """Generate reports for all months in the data with weekly and monthly sheets."""
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
        
        months = self.get_months_in_data()
        
        if not months:
            print("No data found or no valid dates in the dataset.")
            return
        
        print(f"Found {len(months)} months of data: {[str(m) for m in months]}")
        
        for month_period in months:
            print(f"\nProcessing {month_period}...")
            
            # Create workbook for this month
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Filter data for this month
            month_data = self.filter_data_by_month(month_period)
            
            if month_data.empty:
                print(f"No data found for {month_period}")
                continue
            
            # Get weeks in this month
            weeks = self.get_weeks_in_month(month_period)
            
            # Generate weekly sheets
            for week_num, week_period in enumerate(weeks, 1):
                week_data = self.filter_data_by_week(week_period)
                if week_data.empty:
                    continue
                
                # Generate summaries for week sheet
                daily_summary = self.generate_daily_summary(week_data, is_week_sheet=True)
                user_summary = self.generate_user_summary(week_data, f"Weekly")
                
                # Create week title and dates
                start_date = week_period.start_time.strftime('%d/%m/%Y')
                end_date = (week_period.start_time + timedelta(days=4)).strftime('%d/%m/%Y')  # Friday
                week_title = f"Image Enhancement Report for Week{week_num} ({start_date} - {end_date})"
                
                # Create worksheet
                sheet_name = f"Week{week_num}"
                self.create_worksheet(wb, sheet_name, daily_summary, user_summary, 
                                    week_title, f"{start_date} - {end_date}", is_week_sheet=True)
            
            # Generate monthly sheet
            daily_summary = self.generate_daily_summary(month_data, is_week_sheet=False)
            user_summary = self.generate_user_summary(month_data, "Monthly")
            
            # Create month title and dates
            month_name = month_period.strftime('%B %Y')
            start_date = month_period.start_time.strftime('%d/%m/%Y')
            end_date = month_period.end_time.strftime('%d/%m/%Y')
            month_title = f"Image Enhancement Report for Month {month_name} ({start_date} - {end_date})"
            
            # Create monthly worksheet
            self.create_worksheet(wb, "Month", daily_summary, user_summary, 
                                month_title, f"{start_date} - {end_date}", is_week_sheet=False)
            
            # Reorder sheets: Week1, Week2, ..., Month
            sheet_names = list(wb.sheetnames)
            week_sheets = [name for name in sheet_names if name.startswith('Week')]
            week_sheets.sort(key=lambda x: int(x.replace('Week', '')))
            
            # Reorder sheets
            for i, sheet_name in enumerate(week_sheets + ['Month']):
                wb._sheets[i] = wb[sheet_name]
            
            # Create output filename
            month_str = month_period.strftime('%Y_%m_%B')
            excel_filename = f"Image_Enhancement_Report_{month_str}.xlsx"
            excel_path = os.path.join(output_directory, excel_filename)
            
            # Save the workbook
            wb.save(excel_path)
            
            print(f"Report generated for {month_period}: {excel_filename}")
            print(f"  - Created {len(weeks)} weekly sheets and 1 monthly sheet")
        
        print(f"\nAll reports have been generated in the '{output_directory}' directory.")

def main():
    """Main function to run the report generator."""
    # Get the CSV file path from user input or use default
    csv_file = input("Enter the path to your CSV file (or press Enter for 'Image enhancement report raw data.csv'): ").strip()
    
    if not csv_file:
        csv_file = "Image enhancement report raw data.csv"

    if not os.path.exists(csv_file):
        print(f"Error: File '{csv_file}' not found.")
        return
    
    try:
        # Create report generator
        generator = ImageEnhancementReportGenerator(csv_file)
        
        # Generate reports
        generator.generate_reports()
        
        print("\n✅ Report generation completed successfully!")
        
    except Exception as e:
        print(f"❌ An error occurred: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()