# utils/excel_generator.py
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Dict, List, Any
from io import BytesIO
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.column_widths = {
            'A': 12,  # Work Date
            'B': 35,  # User
            'C': 10,  # Total Done
            'D': 8,   # Good
            'E': 12,  # Good Original
            'F': 12,  # Good Enhanced
            'G': 10,  # For Editing
            'H': 8,   # Bad
            'I': 10,  # Rejected
            'J': 12,  # Total Reviewed
            'K': 12,  # Downloaded
            'L': 10,  # Uploaded
            'M': 12,  # Total Edited
        }
    
    def create_report(self, processed_data: Dict) -> BytesIO:
        """Generate Excel report from processed data"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        try:
            # Create monthly sheet first
            self._create_monthly_sheet(wb, processed_data)
            
            # Create weekly sheets
            weekly_summaries = processed_data.get('weekly_summaries', [])
            for i, week_data in enumerate(weekly_summaries, 1):
                self._create_weekly_sheet(wb, week_data, processed_data, i)
            
            # Reorder sheets: Week1, Week2, ..., Month
            self._reorder_sheets(wb)
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error generating Excel report: {str(e)}")
    
    def _create_monthly_sheet(self, wb: Workbook, processed_data: Dict):
        """Create monthly summary sheet"""
        ws = wb.create_sheet(title="Month")
        
        # Get monthly data
        monthly_summaries = processed_data.get('monthly_summaries', [])
        if not monthly_summaries:
            return
        
        month_data = monthly_summaries[0]  # Assuming single month for now
        
        # Set title
        title = f"Image Enhancement Report for Month {month_data.get('month_name', 'Unknown')} ({month_data.get('start_date', '')} - {month_data.get('end_date', '')})"
        ws['A1'] = title
        ws.merge_cells('A1:M1')
        
        # Style title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add section headers
        ws['D2'] = 'REVIEWER'
        ws.merge_cells('D2:J2')
        ws['D2'].font = Font(bold=True)
        ws['D2'].alignment = Alignment(horizontal='center')
        
        ws['K2'] = 'EDITOR'
        ws.merge_cells('K2:M2')
        ws['K2'].font = Font(bold=True)
        ws['K2'].alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ['Work Date', 'User', 'Total Done', 'Good', 'Good Original', 'Good Enhanced',
                  'For Editing', 'Bad', 'Rejected', 'Total Reviewed', 'Downloaded', 'Uploaded', 'Total Edited']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add daily data
        row = 4
        daily_summaries = processed_data.get('daily_summaries', [])
        
        for daily_data in daily_summaries:
            date_str = daily_data.get('date', '')
            weekday = daily_data.get('weekday', '')
            
            # Add user records for this day
            user_records = daily_data.get('user_records', [])
            date_start_row = row
            
            for user_record in user_records:
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=user_record.get('user', ''))
                ws.cell(row=row, column=3, value=user_record.get('total_done', 0))
                ws.cell(row=row, column=4, value=user_record.get('good', 0))
                ws.cell(row=row, column=5, value=user_record.get('good_original', 0))
                ws.cell(row=row, column=6, value=user_record.get('good_enhanced', 0))
                ws.cell(row=row, column=7, value=user_record.get('for_editing', 0))
                ws.cell(row=row, column=8, value=user_record.get('bad', 0))
                ws.cell(row=row, column=9, value=user_record.get('rejected', 0))
                ws.cell(row=row, column=10, value=user_record.get('total_reviewed', 0))
                ws.cell(row=row, column=11, value=user_record.get('downloaded', 0))
                ws.cell(row=row, column=12, value=user_record.get('uploaded', 0))
                ws.cell(row=row, column=13, value=user_record.get('total_edited', 0))
                row += 1
            
            # Merge date cells if multiple users for same date
            if len(user_records) > 1:
                ws.merge_cells(f'A{date_start_row}:A{row-1}')
                ws.cell(row=date_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Add daily total
            daily_totals = daily_data.get('daily_totals', {})
            ws.cell(row=row, column=1, value=f'TOTAL {weekday.upper()}')
            ws.cell(row=row, column=2, value='')
            ws.cell(row=row, column=3, value=daily_totals.get('total_done', 0))
            ws.cell(row=row, column=4, value=daily_totals.get('good', 0))
            ws.cell(row=row, column=5, value=daily_totals.get('good_original', 0))
            ws.cell(row=row, column=6, value=daily_totals.get('good_enhanced', 0))
            ws.cell(row=row, column=7, value=daily_totals.get('for_editing', 0))
            ws.cell(row=row, column=8, value=daily_totals.get('bad', 0))
            ws.cell(row=row, column=9, value=daily_totals.get('rejected', 0))
            ws.cell(row=row, column=10, value=daily_totals.get('total_reviewed', 0))
            ws.cell(row=row, column=11, value=daily_totals.get('downloaded', 0))
            ws.cell(row=row, column=12, value=daily_totals.get('uploaded', 0))
            ws.cell(row=row, column=13, value=daily_totals.get('total_edited', 0))
            
            # Bold the total row
            for col in range(1, 14):
                ws.cell(row=row, column=col).font = Font(bold=True)
            
            row += 1
        
        # Add monthly total
        if monthly_summaries:
            month_total = monthly_summaries[0]
            ws.cell(row=row, column=1, value='MONTHLY TOTAL')
            ws.cell(row=row, column=2, value='')
            ws.cell(row=row, column=3, value=month_total.get('total_done', 0))
            ws.cell(row=row, column=4, value=month_total.get('good_count', 0))
            ws.cell(row=row, column=10, value=month_total.get('total_reviewed', 0))
            ws.cell(row=row, column=13, value=month_total.get('total_edited', 0))
            
            # Bold the monthly total
            for col in range(1, 14):
                ws.cell(row=row, column=col).font = Font(bold=True)
        
        # Add user summary section
        self._add_user_summary_section(ws, processed_data, row + 3)
        
        # Apply formatting
        self._apply_sheet_formatting(ws)
    
    def _create_weekly_sheet(self, wb: Workbook, week_data: Dict, processed_data: Dict, week_num: int):
        """Create weekly sheet"""
        sheet_name = f"Week{week_num}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Set title
        title = f"Image Enhancement Report for Week{week_num} ({week_data.get('start_date', '')} - {week_data.get('end_date', '')})"
        ws['A1'] = title
        ws.merge_cells('A1:M1')
        
        # Style title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add section headers
        ws['D2'] = 'REVIEWER'
        ws.merge_cells('D2:J2')
        ws['D2'].font = Font(bold=True)
        ws['D2'].alignment = Alignment(horizontal='center')
        
        ws['K2'] = 'EDITOR'
        ws.merge_cells('K2:M2')
        ws['K2'].font = Font(bold=True)
        ws['K2'].alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ['Work Date', 'User', 'Total Done', 'Good', 'Good Original', 'Good Enhanced',
                  'For Editing', 'Bad', 'Rejected', 'Total Reviewed', 'Downloaded', 'Uploaded', 'Total Edited']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get daily data for this week
        row = 4
        daily_breakdown = week_data.get('daily_breakdown', [])
        
        # Filter daily summaries for this week's dates
        week_start = week_data.get('start_date', '')
        week_end = week_data.get('end_date', '')
        
        daily_summaries = processed_data.get('daily_summaries', [])
        week_dailies = [d for d in daily_summaries if week_start <= d.get('date', '') <= week_end]
        
        for daily_data in week_dailies:
            date_str = daily_data.get('date', '')
            weekday = daily_data.get('weekday', '')
            
            # Add user records
            user_records = daily_data.get('user_records', [])
            date_start_row = row
            
            for user_record in user_records:
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=user_record.get('user', ''))
                ws.cell(row=row, column=3, value=user_record.get('total_done', 0))
                ws.cell(row=row, column=4, value=user_record.get('good', 0))
                ws.cell(row=row, column=5, value=user_record.get('good_original', 0))
                ws.cell(row=row, column=6, value=user_record.get('good_enhanced', 0))
                ws.cell(row=row, column=7, value=user_record.get('for_editing', 0))
                ws.cell(row=row, column=8, value=user_record.get('bad', 0))
                ws.cell(row=row, column=9, value=user_record.get('rejected', 0))
                ws.cell(row=row, column=10, value=user_record.get('total_reviewed', 0))
                ws.cell(row=row, column=11, value=user_record.get('downloaded', 0))
                ws.cell(row=row, column=12, value=user_record.get('uploaded', 0))
                ws.cell(row=row, column=13, value=user_record.get('total_edited', 0))
                row += 1
            
            # Merge date cells
            if len(user_records) > 1:
                ws.merge_cells(f'A{date_start_row}:A{row-1}')
                ws.cell(row=date_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

            # Add daily total
            daily_totals = daily_data.get('daily_totals', {})
            ws.cell(row=row, column=1, value=f'TOTAL {weekday.upper()}')
            ws.cell(row=row, column=3, value=daily_totals.get('total_done', 0))
            ws.cell(row=row, column=10, value=daily_totals.get('total_reviewed', 0))
            ws.cell(row=row, column=13, value=daily_totals.get('total_edited', 0))
            
            for col in range(1, 14):
                ws.cell(row=row, column=col).font = Font(bold=True)
            
            row += 1
        
        # Add weekly total
        ws.cell(row=row, column=1, value='WEEKLY TOTAL')
        ws.cell(row=row, column=3, value=week_data.get('total_done', 0))
        ws.cell(row=row, column=10, value=week_data.get('total_reviewed', 0))
        ws.cell(row=row, column=13, value=week_data.get('total_edited', 0))
        
        for col in range(1, 14):
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        # Add user summary section
        self._add_user_summary_section(ws, processed_data, row + 3, is_weekly=True, week_num=week_num)
        
        # Apply formatting
        self._apply_sheet_formatting(ws)

    def _add_user_summary_section(self, ws, processed_data: Dict, start_row: int, is_weekly: bool = False, week_num: int = 0):
        """Add user summary section to a sheet"""
        ws.cell(row=start_row, column=2, value='USER SUMMARY').font = Font(bold=True, size=12)
        
        headers = ['User', 'Total Done', 'Good', 'Good Original', 'Good Enhanced', 'For Editing', 
                   'Bad', 'Rejected', 'Total Reviewed', 'Downloaded', 'Uploaded', 'Total Edited']
        
        for col, header in enumerate(headers, 2):
            ws.cell(row=start_row + 1, column=col, value=header).font = Font(bold=True)

        user_summaries = processed_data.get('user_summaries', [])
        
        row = start_row + 2
        for user_data in user_summaries:
            ws.cell(row=row, column=2, value=user_data.get('user', ''))
            ws.cell(row=row, column=3, value=user_data.get('total_done', 0))
            ws.cell(row=row, column=4, value=user_data.get('good_count', 0))
            ws.cell(row=row, column=8, value=user_data.get('bad_count', 0))
            ws.cell(row=row, column=9, value=user_data.get('rejected_count', 0))
            ws.cell(row=row, column=10, value=user_data.get('total_reviewed', 0))
            ws.cell(row=row, column=13, value=user_data.get('total_edited', 0))
            row += 1
            
        # Add total row
        total_row = row
        total_label = f'WEEK {week_num} TOTAL' if is_weekly else 'MONTHLY TOTAL'
        ws.cell(row=total_row, column=2, value=total_label).font = Font(bold=True)
        
        # Sum up user totals for the grand total
        df = pd.DataFrame(user_summaries)
        ws.cell(row=total_row, column=3, value=df['total_done'].sum()).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=df['good_count'].sum()).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=df['bad_count'].sum()).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=df['rejected_count'].sum()).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=df['total_reviewed'].sum()).font = Font(bold=True)
        ws.cell(row=total_row, column=13, value=df['total_edited'].sum()).font = Font(bold=True)

    def _apply_sheet_formatting(self, ws):
        """Apply borders, column widths, and alignment"""
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.thin_border
        
        for col_letter, width in self.column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _reorder_sheets(self, wb: Workbook):
        """Reorder sheets to have weekly sheets first, then the monthly sheet."""
        sheet_names = wb.sheetnames
        
        # Separate weekly sheets from others
        weekly_sheets = sorted([name for name in sheet_names if name.startswith('Week')])
        other_sheets = [name for name in sheet_names if not name.startswith('Week')]
        
        # New order: Week1, Week2, ..., Month
        new_order = weekly_sheets + other_sheets
        
        # Reorder
        wb._sheets = [wb[name] for name in new_order]
